import os
import xlrd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re


# create folder new_folder
# open excel file and parse through lines
#   take file names from C NAME index 1
#   take folder name from
#       LIS_CTRY index 3
#       C NAME index 1
# open folder matching folder name without capitalization
# search through file names for match
# if match:
#   copy file into new_folder
#   rename file as CompanyX_Year
#   highlight line in excel file
# else:
#   move to next line


def rename_file(basepath, row_array, row_count, sheet, _inner, index):
    if index == 0:
        return 0
    final_filename = ''
    year_array = [2014, 2015, 2016, 2017]
    try:
        year_array.index(int(final_filename.join(re.findall("[0-9]", os.listdir(basepath)[index]))))
        if os.path.isfile(os.path.join(basepath, os.listdir(basepath)[index])):
            try:
                year_array.index(int(sheet.cell_value(row_count, 4)))
                try:
                    row_array.index(row_count)
                except ValueError:
                    row_array.append(row_count)
            except ValueError:
                pass

            src = basepath + '/' + os.listdir(basepath)[index]
            dst = './new_folder'
            shutil.copy2(src, dst)
            s = ""
            folder_name_inner = _inner.replace(' ', '').replace('-', '')

            os.rename(dst + '/' + os.listdir(basepath)[index], dst + '/' + folder_name_inner + '_' + s.join(
                re.findall("[0-9]", os.listdir(basepath)[index])) + '.pdf')
    except ValueError:
        pass
    rename_file(basepath, row_array, row_count, sheet, _inner, index - 1)


def find_name(directory, count, folder_name_outer, folder_name_split, index):
    if index == 0:
        return False
    if folder_name_split[index].upper() == folder_name_outer.upper() or directory[count].find(folder_name_split[index]):
        return True
    else:
        find_name(directory, count, folder_name_outer, folder_name_split, index - 1)


def calc_rows(sheet, _path, row_array, row_count):
    if row_count == 0:
        return 0
    folder_name_outer = sheet.cell_value(row_count, 3)
    folder_name_inner = sheet.cell_value(row_count, 1)

    if _path.capitalize() == folder_name_outer.capitalize():
        if len(os.listdir('./' + _path)) != 0:
            directory = os.listdir('./' + _path)
            try:
                directory.index(folder_name_inner)
            except ValueError:
                if directory[0].istitle():
                    folder_name_inner = folder_name_inner.capitalize()
                elif directory[0].isupper():
                    folder_name_inner = folder_name_inner.upper()
                elif directory[0].islower():
                    folder_name_inner = folder_name_inner.lower()
            basepath = './' + folder_name_outer + '/' + folder_name_inner
            if len(os.listdir(basepath)) != 0:
                for count in range(len(directory)):
                    if find_name(directory, count, folder_name_outer, folder_name_inner.split(' '),
                                 len(folder_name_inner.split(' '))):
                        basepath = './' + folder_name_outer + '/' + directory[row_count]
            rename_file(basepath, row_array, row_count, sheet, folder_name_inner, len(os.listdir(basepath)))

        else:
            print('the folder was not found A')
    calc_rows(sheet, _path, row_array, row_count - 1)
    format_cells(row_array)


def renaming():
    try:
        os.mkdir('./new_folder')
    except FileExistsError:
        print('folder already exists')

    loc = './excel.xlsx'
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_name('Sheet1')
    check_path(os.listdir('.'), len(os.listdir('.')), sheet, sheet.nrows)


def check_path(_dir, index, sheet, nrows):
    row_array = []
    if os.path.isdir('./' + _dir[index]):
        if _dir[index] == '.idea' or _dir[index] == 'new_folder':
            return 0
        calc_rows(sheet, _dir[index], row_array, nrows)
    check_path(_dir, index - 1, sheet, nrows)


def enumerate_cells(row_array, index, wksht, alpha):
    col = 0
    while col <= wksht.max_column:
        cell = wksht[alpha[col] + str(row_array[index])]
        cell.fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
        col = col + 1
    enumerate_cells(row_array, index - 1)


def format_cells(row_array):
    file = 'excel.xlsx'
    wkbk = load_workbook(filename=file)
    # Enumerate the cells in the second row
    enumerate_cells(row_array, len(row_array), wkbk['Sheet1'], ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'])
    wkbk.save(filename=file)
